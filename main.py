import requests
import schedule
import time
from flask import Flask, request, redirect, jsonify
from threading import Thread
import logging
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from datetime import datetime, timedelta
import pymysql
from collections import deque
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from cookie import login_and_fetch_cookie  # 从cookie.py导入方法
try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available, Excel formatting will be limited")

CHROME_DRIVER_PATH = "/opt/homebrew/bin/chromedriver" #

app = Flask(__name__)

# Configure logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s - %(message)s')

# Global variable to store the cookie
cookie_value = 'wk_cookie2=167f0c4d672ac968e0a93b99b0d0321b; __itrace_wid=95ae536e-98c9-4b4d-80b3-e28c2a186e6a; cna=VWfNIPe5zgoCASvziD7zfIGi; env_bak=FM%2BgywHD7Unoz8SbIc%2FnUBfFQSvsP66Kkma08fliLj6Y; __wpkreporterwid_=08407be1-a5ed-4adc-3102-f98773652212; t_alimama=ee2cffdec21995b8105c3f08a73ab6c9; lgc=; cookie2=1e5d78ccd2b97d0d1da082d51dd217ce; t=aad3d47a3693faf2be2983e66f9366c7; sn=; _tb_token_=7384e3d9e335a; XSRF-TOKEN=cb12569d-1278-439d-adef-7ace59411b33; cookie3_bak=1e5d78ccd2b97d0d1da082d51dd217ce; login=true; cancelledSubSites=empty; xlly_s=1; dnk=%5Cu5929%5Cu6D25%5Cu6D69%5Cu777F%5Cu4FE1%5Cu606F%5Cu79D1%5Cu6280%5Cu6709%5Cu9650%5Cu516C%5Cu53F8; tracknick=%5Cu5929%5Cu6D25%5Cu6D69%5Cu777F%5Cu4FE1%5Cu606F%5Cu79D1%5Cu6280%5Cu6709%5Cu9650%5Cu516C%5Cu53F8; lid=%E5%A4%A9%E6%B4%A5%E6%B5%A9%E7%9D%BF%E4%BF%A1%E6%81%AF%E7%A7%91%E6%8A%80%E6%9C%89%E9%99%90%E5%85%AC%E5%8F%B8; wk_unb=UUpgTs0%2BqS386PihCA%3D%3D; _l_g_=Ug%3D%3D; cookie1=U7ZvqNzFjt0WHllf4lbxAgQ5Wl4LJjp%2Fl6HVIQ%2FXkiU%3D; sg=%E5%8F%B85e; uc1=cookie14=UoYagke%2BQnvi0A%3D%3D&pas=0&cookie15=V32FPkk%2Fw0dUvg%3D%3D&existShop=false&cookie16=UtASsssmPlP%2Ff1IHDsDaPRu%2BPw%3D%3D&cookie21=VT5L2FSpdiBh; havana_lgc_exp=1780898242562; unb=2218492598755; cookie17=UUpgTs0%2BqS386PihCA%3D%3D; _nk_=%5Cu5929%5Cu6D25%5Cu6D69%5Cu777F%5Cu4FE1%5Cu606F%5Cu79D1%5Cu6280%5Cu6709%5Cu9650%5Cu516C%5Cu53F8; sgcookie=E100Iwg5vQePW%2FuvoH870NOoXXaNTJ0n3ETkubPIWgdgirmz3UJSBvaSCd7DDVlbUy1zB%2BSIN9KnVy%2B5Hh5yVVEzOPAkx3QSfJAoimBHTiaJr4A%3D; csg=5ce22315; cookie3_bak_exp=1750053442563; tfstk=gGIK-W2eFRHplBjpI9ziZYhTGPegiPXFQ6WjqQAnP1CO3_LoZTaUeuCGwe2FFW5R2C1ytBxuLUtR36dHrHtk2bCcUBgHZw2JNC6XEJyzUudR5tbHEw2z2QKFjwmktW-RFsx8iSq0m9Wev3N0iYRfPmteeLMBFbc_fLtJd1faKXWe43NiI0a0D95FfMKVN3w9fLprP3OIVfw9nL-SAUiI5fOw139WdB66Cpv-VDO5NOw9UCOWV31WfRpr_w2p_v9K2ZAxuReKeGiSVGpp5PXB1ETYEpTpG9OTG_SMpQRfdCnSVIBNNTW5N7nkcZ5O2E1z2c-hHNt6pNeS5BTAR6vNwonBOGBCABjQsDACYO_cq1wSVQQ1lEAC75ryRZ5hlFIg9cRf5OBkWNyr-Qb2F_8VqSoeOT6V0ZxY2YpOkOtO4BjcD96siIpoRRetz48BQ0A3ih05JwJWBI2BR4uyrdooJgJjz48tmdd0IOurzF0c.; isg=BMrKpYwglklzYxrv_SRNU794G7Zsu04VvzPZEVQDopw0B2vBPE6OJgn1E3Pb98at'

# 默认的广告位列表
ad_slots = [
    "mm_1902210064_2348000105_111662900182",
    "mm_1902210064_2348000105_111663100185",
    "mm_1902210064_2348000105_113763000348",
    "mm_1902210064_2348000105_113765750160",
    "mm_3447365382_2749500311_114553400117",
    "mm_3447365382_2749500311_114552150188",
    "mm_1873810155_2320450209_111384500336",
    "mm_1873810155_2320450209_111953150429",
    "mm_1873810155_2320450209_111953700467",
    "mm_1873810155_2320450209_115792250490",
    "mm_1873810155_2320450209_115890350442",
    "mm_1873810155_2320450209_115894200243",
    "mm_1873810155_2320450209_115797000138",
    "mm_1873810155_2320450209_111388400121",
    "mm_1861850082_2364600045_115796950045",
    "mm_1861850082_2364600045_111952850176",
    "mm_1861850082_2364600045_111952350174",
    "mm_1861850082_2364600045_111484350168",
    "mm_1562690005_2184850029_111650700372",
    "mm_1562690005_2184850029_111651000378",
    "mm_1562690005_2184850029_112164850011",
    "mm_1562690005_2184850029_112159050469",
    "mm_6899417631_3131950139_115762350460",
    "mm_6899417631_3131950139_115765600056",
    "mm_6899417631_3131950139_115768000140",
    "mm_6899417631_3131950139_115764850334",
    "mm_6899417631_3131950139_115840550019",
    "mm_6899417631_3131950139_115834100467",
    "mm_2279650033_2540650473_111838750084",
    "mm_2279650033_2540650473_111836000276",
    "mm_2279650033_2540650473_111835400327",
    "mm_2279650033_2540650473_111835250296",
    "mm_2279650033_2540650473_111835500278",
    "mm_2279650033_2540650473_111835750274",
    "mm_2279650033_2540650473_114534000022",
    "mm_2279650033_2540650473_114532100161",
]

ad_slots_up = []

# MySQL database configuration
# DB_CONFIG = {
#     'host': '172.16.3.12',
#     'user': 'adx',
#     'password': 'CaPn1jxidkkE5',
#     'database': 'release_atd',
#     'charset': 'utf8mb4'
# }

DB_CONFIG = {
    'host': '127.0.0.1',
    'user': 'root',
    'password': '123456',
    'database': 'release_atd',
    'charset': 'utf8mb4'
}

# Establish a global database connection at project startup
connection = pymysql.connect(**DB_CONFIG)

# Ensure the DB connection is alive or reopen if dropped
def get_connection():
    global connection
    try:
        connection.ping(reconnect=True)
    except Exception:
        connection = pymysql.connect(**DB_CONFIG)
    return connection

# Function to insert data into the MySQL table
def insert_data(ds, pid, adzone_name, qingqiupv, active_ratio_df, tanx_effect_pv, tanx_clk, dongfeng_ef):
    try:
        conn = get_connection()
        with conn.cursor() as cursor:
            sql = '''
            INSERT INTO tanx_monitor (ds, pid, adzone_name, qingqiupv, active_ratio_df, tanx_effect_pv, tanx_clk, dongfeng_ef, create_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, UNIX_TIMESTAMP())
            ON DUPLICATE KEY UPDATE
                adzone_name = VALUES(adzone_name),
                qingqiupv = VALUES(qingqiupv),
                active_ratio_df = VALUES(active_ratio_df),
                tanx_effect_pv = VALUES(tanx_effect_pv),
                tanx_clk = VALUES(tanx_clk),
                dongfeng_ef = VALUES(dongfeng_ef)
            '''
            cursor.execute(sql, (ds, pid, adzone_name, qingqiupv, active_ratio_df, tanx_effect_pv, tanx_clk, dongfeng_ef))
        connection.commit()
    except Exception as e:
        logging.error(f"Failed to insert data for pid {pid}: {e}")
        print(f"Failed to insert data for pid {pid}: {e}")


@app.route('/')
def cookie_input_page():
    try:
        # 使用绝对路径加载 HTML 文件
        html_path = './cookie_input.html'
        with open(html_path, 'r', encoding='utf-8') as file:
            html_content = file.read()
        # 将广告位列表插入到 <textarea> 中
        ad_slots_text = "\n".join(ad_slots)
        html_content = html_content.replace('<textarea id="ad_slots" name="ad_slots" rows="5" required></textarea>', f'<textarea id="ad_slots" name="ad_slots" rows="5" required>{ad_slots_text}</textarea>')
        return html_content, 200, {'Content-Type': 'text/html'}
    except FileNotFoundError:
        logging.error("cookie_input.html not found")
        return "Error: cookie_input.html not found", 404
    

@app.route('/update_cookie', methods=['POST'])
def update_cookie():
    global cookie_value
    cookie_value = request.form['cookie']
    return "Cookie 已更新成功！"


@app.route('/update_ad_slots', methods=['POST'])
def update_ad_slots():
    global ad_slots
    global ad_slots_up
    # 去掉传递进来的数据中的 \r
    ad_slots = [slot.replace('\r', '') for slot in request.form['ad_slots'].split('\n')]
    if len(ad_slots_up) == 0:
        ad_slots_up = ad_slots.copy()
    logging.info(f"广告位列表已更新: {ad_slots}")
    return "广告位列表已更新成功！"

# Execute fetch_data once at startup
def fetch_data():
    status_record = {"execution_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "status": "执行中"}
    #print(f"Fetching data at {status_record['execution_time']} with cookie: {cookie_value}")
    try:
        url = 'https://tanx.alimama.com/api/media/debug/report/getReport.htm'
        headers = {
            'Cookie': cookie_value,
            'Content-Type': 'application/json'
        }
        # Use configurable days back instead of fixed yesterday
        query_date = (datetime.now() - timedelta(days=query_days_back)).strftime('%Y-%m-%d')
        print(f"Fetching data for {query_date} (前{query_days_back}天)")
    
        for pid in ad_slots:
            data = {
                "ds": query_date,
                "mediaClick": "1",
                "mediaCost": "1",
                "mediaPV": "1",
                "pid": pid,
                "type": "rtb"
            }
            response = requests.post(url, headers=headers, json=data)
            #print(f"Fetching data for url:{url} pid: {pid} with data: {data} response data: {response}")
            try:
                response_data = response.json()
                if response_data.get("info", {}).get("errorCode") == "user_not_login":
                    logging.info("User not logged in, please update cookie.")
                    update_cookie_task()  # 自动更新cookie
                tanx_monitor_param_list = response_data.get("data", {}).get("clickMonitorParamList", [])
                # 修改条件判断，检查列表是否为空
                if len(tanx_monitor_param_list) == 0:
                    logging.error(f"No data found for pid {pid} on {query_date}")
                    print(f"No data found for pid {pid} on {query_date}")
                    continue
                # 将 activeRatioDf 转换为百分数格式
                for item in tanx_monitor_param_list:
                    active_ratio_df = item.get("activeRatioDf") if item.get("activeRatioDf") not in (None, "") else "0"
                    ratio = float(active_ratio_df) * 100
                    active_ratio_df_percent = str(f"{ratio:.2f}%")
                    insert_data(
                        item.get("ds"),
                        item.get("pid"),
                        item.get("adzoneName"),
                        item.get("qingqiupv"),
                        active_ratio_df_percent,
                        item.get("tanxEffectPv"),
                        item.get("tanxClk"),
                        item.get("dongfengEf")
                    )
                    logging.info(f"Stored Data for pid {pid}: {item}")
                    print(f"Stored Data for pid: {pid}")
            except Exception as e:
                logging.error(f"Error processing response for pid {pid}: {e}")
                print(f"Error processing response for pid {pid}: {e}")
        status_record["status"] = "成功"
    except Exception as e:
        status_record["status"] = f"失败: {e}"
    finally:
        # 修复 scheduler_status 的错误操作
        # 删除错误的字典键操作
        # scheduler_status["last_execution"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 确保 fetch_data 函数正确追加 status_record
        scheduler_status.append(status_record)

@app.route('/fetch_data', methods=['POST'])
def fetch_data_button():
    fetch_data()
    query_and_export_data()
    return "抓包调用成功！"

def run_flask():
    app.run(port=5003)


# 全局变量存储定时任务状态
scheduler_status = deque(maxlen=10)  # 存储最近十次的执行记录

@app.route('/scheduler_status', methods=['GET'])
def get_scheduler_status():
    return list(scheduler_status)

def extract_media(adzone_name):
    if pd.isna(adzone_name):
        return '未知'
    for key in ['佳投', '有境', '新数', '快友', '多盟', '浩睿', '美数']:
        if key in str(adzone_name):
            return key
    return '其他'

def add_media_statistics(df):
    # 临时添加媒体列用于分组，但不包含在最终结果中
    df_temp = df.copy()
    df_temp['媒体'] = df_temp['广告位名称'].apply(extract_media)
    result_rows = []
    
    # 按日期和媒体分组
    for date in df_temp['日期'].unique():
        formatted_date = pd.to_datetime(str(date), format='%Y%m%d').strftime('%Y/%m/%d')
        date_df = df_temp[df_temp['日期'] == date]
        
        for media in date_df['媒体'].unique():
            media_date_df = date_df[date_df['媒体'] == media]
            
            # 添加该媒体当天的所有数据行（不包含媒体列）
            for _, row in media_date_df.iterrows():
                row_dict = row.to_dict()
                # 删除媒体列
                if '媒体' in row_dict:
                    del row_dict['媒体']
                if '日期' in row_dict:
                    row_dict['日期'] = formatted_date
                result_rows.append(row_dict)
            
            # 添加该媒体当天的统计行
            stats_row = {
                '日期': formatted_date,
                '广告位': '统计总数',
                '广告位名称': f'{media}',
                'tanx有效请求': int(media_date_df['tanx有效请求'].apply(pd.to_numeric, errors='coerce').sum()),
                '东风手淘换端率-同步点击': f"{media_date_df['东风手淘换端率-同步点击'].apply(lambda x: float(str(x).replace('%','')) if '%' in str(x) and str(x).replace('%','').replace('.','',1).isdigit() else None).mean():.2f}%",
                'TANX曝光数': int(media_date_df['TANX曝光数'].apply(pd.to_numeric, errors='coerce').sum()),
                'TANX点击数': int(media_date_df['TANX点击数'].apply(pd.to_numeric, errors='coerce').sum()),
                'TANX预估收益': float(media_date_df['TANX预估收益'].apply(pd.to_numeric, errors='coerce').sum())
            }
            logging.info(f"Date: {date}, Media: {media}, Stats: {stats_row}")
            result_rows.append(stats_row)
            
        # 每天结束后添加空行分隔（不包含媒体列）
        empty_row = {col: '' for col in df.columns}
        result_rows.append(empty_row)

    logging.info(f"Final DataFrame with daily statistics:\n{pd.DataFrame(result_rows)}")
    result_df = pd.DataFrame(result_rows)
    return result_df[df.columns]  # 保持原有列顺序（不包含媒体列）

def create_summary_sheet(df_with_stats):
    """从详细数据中提取所有统计汇总行创建汇总表"""
    # 筛选出所有统计汇总行（广告位为"统计总数"的行）
    summary_rows = df_with_stats[df_with_stats['广告位'] == '统计总数'].copy()
    
    logging.info(f"Found {len(summary_rows)} summary rows from detailed data")
    logging.info(f"Summary rows shape: {summary_rows.shape}")
    
    if summary_rows.empty:
        logging.warning("No summary rows found in detailed data")
        return pd.DataFrame()
    
    # 重置索引
    summary_rows.reset_index(drop=True, inplace=True)
    
    logging.info(f"Created summary sheet with {len(summary_rows)} rows")
    return summary_rows

def query_and_export_data():
    try:
        # Query the database for the last 30 days of data
        query = '''
        SELECT ds, pid, adzone_name, qingqiupv, active_ratio_df, tanx_effect_pv, tanx_clk, dongfeng_ef
        FROM tanx_monitor
        WHERE ds >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
        '''
        conn = get_connection()
        with conn.cursor() as cursor:
            cursor.execute(query)
            result = cursor.fetchall()

        # Convert the result to a pandas DataFrame
        columns = ['日期', '广告位', '广告位名称', 'tanx有效请求', '东风手淘换端率-同步点击', 'TANX曝光数', 'TANX点击数', 'TANX预估收益']
        df = pd.DataFrame(result, columns=columns)
        
        # 添加媒体统计到原始数据
        df_with_stats = add_media_statistics(df)
        
        # 从详细数据中提取汇总行创建汇总表
        summary_df = create_summary_sheet(df_with_stats)
        logging.info(f"Summary DataFrame shape: {summary_df.shape}")
        logging.info(f"Summary DataFrame empty: {summary_df.empty}")

        # Export to Excel with multiple sheets
        #file_path = '/tmp/tanx_data.xlsx'
        file_path = 'D:\\tanx\\tanx_data.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 写入详细数据到第一个sheet
            df_with_stats.to_excel(writer, sheet_name='详细数据', index=False)
            
            # 写入汇总数据到第二个sheet
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
                logging.info("Summary sheet created successfully")
            else:
                logging.warning("Summary sheet not created - no summary data found")
        
        # 调整Excel列宽（对两个sheet都进行调整）
        adjust_excel_column_width(file_path)
        
        logging.info(f"Data exported to {file_path} with detailed data and summary statistics")

        # Send the Excel file via email
        send_email(file_path)

    except Exception as e:
        logging.error(f"Error querying or exporting data: {e}")

def adjust_excel_column_width(file_path):
    """调整Excel列宽，使表格更易读"""
    try:
        if not OPENPYXL_AVAILABLE:
            logging.warning("openpyxl not available, skipping column width adjustment")
            return
            
        # 加载工作簿
        wb = load_workbook(file_path)
        
        # 定义每列的最小宽度
        column_widths = {
            'A': 12,  # 日期
            'B': 25,  # 广告位
            'C': 30,  # 广告位名称
            'D': 15,  # tanx有效请求
            'E': 25,  # 东风手淘换端率-同步点击
            'F': 15,  # TANX曝光数
            'G': 15,  # TANX点击数
            'H': 15,  # TANX预估收益
        }
        
        # 对所有工作表进行列宽调整
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置列宽
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # 自动调整列宽（基于内容）
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 设置调整后的宽度，但不超过50个字符
                adjusted_width = min(max_length + 2, 50)
                # 确保不小于预设的最小宽度
                final_width = max(adjusted_width, column_widths.get(column_letter, 10))
                ws.column_dimensions[column_letter].width = final_width
        
        # 保存文件
        wb.save(file_path)
        logging.info("Excel列宽调整完成（所有工作表）")
        
    except Exception as e:
        logging.error(f"调整Excel列宽失败: {e}")

# Global variable to store email recipients
default_email_recipients = ['chemanyu@admate.cn','zhangwenjing@admate.cn','xuzhongwang@admate.cn','fanang@admate.cn']
#default_email_recipients = ['chemanyu@admate.cn', 'fanang@admate.cn']
email_recipients = default_email_recipients.copy()


# Global variable to store query days back
default_days_back = 1
query_days_back = default_days_back

def send_email(file_path):
    try:
        # Email configuration
        sender_email = "monitor@admate.cn"
        sender_password = "Wut49622"
        subject = "Tanx Data Export"

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(email_recipients)
        msg['Subject'] = subject

        # Email body
        print(f"Sending email to: {email_recipients}")
        body = "Please find the attached Excel file containing the Tanx data for the last 30 days."
        msg.attach(MIMEText(body, 'plain'))

        # Attach the Excel file
        with open(file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={file_path.split("/")[-1]}')
        msg.attach(part)

        # Send the email
        with smtplib.SMTP('smtp.partner.outlook.cn', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, email_recipients, msg.as_string())

        logging.info(f"Email sent to {email_recipients}")

    except Exception as e:
        logging.error(f"Error sending email: {e}")

# 定时更新cookie的函数
def update_cookie_task():
    global cookie_value
    new_cookie = login_and_fetch_cookie()
    if new_cookie:
        cookie_value = new_cookie
        logging.info("自动更新Cookie成功")
    else:
        logging.error("自动更新Cookie失败")


@app.route('/get_email_recipients', methods=['GET'])
def get_email_recipients():
    return jsonify({
        'current': email_recipients,
        'default': default_email_recipients
    })

@app.route('/update_email_recipients', methods=['POST'])
def update_email_recipients():
    global email_recipients
    try:
        # 获取提交的邮件列表，按行分割并过滤空行
        email_text = request.form.get('email_recipients', '')
        new_emails = [email.strip() for email in email_text.split('\n') if email.strip()]
        
        # 去重并过滤有效邮箱
        valid_emails = []
        for email in new_emails:
            if '@' in email and '.' in email:  # 简单的邮箱格式验证
                if email not in valid_emails:  # 去重
                    valid_emails.append(email)
        
        if valid_emails:
            email_recipients = valid_emails
            logging.info(f"邮件收件人列表已更新: {email_recipients}")
            return "邮件收件人列表已更新成功！"
        else:
            return "错误：请输入至少一个有效的邮箱地址！", 400
            
    except Exception as e:
        logging.error(f"更新邮件收件人列表失败: {e}")
        return f"更新失败: {str(e)}", 500

@app.route('/reset_email_recipients', methods=['POST'])
def reset_email_recipients():
    global email_recipients
    try:
        email_recipients = default_email_recipients.copy()
        logging.info(f"邮件收件人列表已重置为默认值: {email_recipients}")
        return "邮件收件人列表已重置为默认值！"
    except Exception as e:
        logging.error(f"重置邮件收件人列表失败: {e}")
        return f"重置失败: {str(e)}", 500


@app.route('/get_query_days', methods=['GET'])
def get_query_days():
    return jsonify({
        'current': query_days_back,
        'default': default_days_back
    })

@app.route('/update_query_days', methods=['POST'])
def update_query_days():
    global query_days_back
    try:
        days = request.form.get('query_days', '')
        if not days.isdigit():
            return "错误：请输入有效的数字！", 400
        
        days_int = int(days)
        if days_int < 1 or days_int > 30:  # 限制查询范围
            return "错误：查询天数必须在1-30天之间！", 400
        
        query_days_back = days_int
        logging.info(f"查询天数已更新为前{query_days_back}天")
        return f"查询天数已更新为前{query_days_back}天！"
        
    except Exception as e:
        logging.error(f"更新查询天数失败: {e}")
        return f"更新失败: {str(e)}", 500

@app.route('/reset_query_days', methods=['POST'])
def reset_query_days():
    global query_days_back
    try:
        query_days_back = default_days_back
        logging.info(f"查询天数已重置为前{query_days_back}天")
        return f"查询天数已重置为前{query_days_back}天！"
    except Exception as e:
        logging.error(f"重置查询天数失败: {e}")
        return f"重置失败: {str(e)}", 500
    

# 设置定时任务
schedule.every().day.at("12:15").do(fetch_data)  # 每天抓取数据
schedule.every().day.at("12:30").do(query_and_export_data)  # 每天导出数据


flask_thread = Thread(target=run_flask)
flask_thread.start()

# 添加调试日志以确认定时任务是否正常运行
try:
    print("Scheduler started. Waiting for tasks...")
    while True:
        schedule.run_pending()
        print("Pending tasks executed.")  # 调试日志
        time.sleep(60)
except KeyboardInterrupt:
    print("Program terminated by user.")
    logging.info("Program terminated by user.")